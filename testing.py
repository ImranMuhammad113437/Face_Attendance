import tkinter as tk
from tkinter import ttk, Listbox
from tkinter import Frame, Canvas, Scrollbar, Label, messagebox, StringVar, BooleanVar, Entry, Listbox
from tkinter import LabelFrame, Label, Entry, Button, StringVar
from tkinter import *
from PIL import Image, ImageTk
import admit_interface
import mysql.connector  # Import your database library
from tkinter import Label, Frame, messagebox  # Ensure required Tkinter classes are imported
import mysql.connector
from tkinter import messagebox  # For showing messages in case of errors
from fpdf import FPDF
from pdf2image import convert_from_path
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from datetime import datetime  # Make sure this module is in the same directory or in your PYTHONPATH
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import calendar
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import os
from tkinter import Frame, Label, Tk
from PIL import Image, ImageDraw, ImageFont, ImageTk

from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.colors import gray
from reportlab.pdfgen import canvas
from reportlab.lib import colors

from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import os

from reportlab.pdfgen import canvas
from tkinter import messagebox
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from tkinter import messagebox
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from tkinter import messagebox
import os
import openpyxl
from tkinter import messagebox


class Report_Generater:
    def __init__(self, root, username):
        self.root = root
        self.username = username  # Assign username to the class instance
        self.root.geometry("1024x590+0+0")
        self.root.title("AttendNow")

        # ------------------------------------- Background Image ------------------------------------------------------
        background_img_main = Image.open(r"Image\Background.png")
        background_img_main = background_img_main.resize((1024, 590), Image.Resampling.LANCZOS)
        self.photo_background_img_main = ImageTk.PhotoImage(background_img_main)
        background_img_main_position = Label(self.root, image=self.photo_background_img_main)
        background_img_main_position.place(x=0, y=0, width=1024, height=590)
        # -------------------------------------------------------------------------------------------------------------

        # ------------------------------------- LogoTitle Image -------------------------------------------------------
        left_title = Image.open(r"Image\LogoTitle_Left Top.png")
        self.photoleft_title = ImageTk.PhotoImage(left_title)
        left_title_position = Label(self.root, image=self.photoleft_title)
        left_title_position.place(x=0, y=0, width=163, height=60)

        # Back Button
        back_button = Button(self.root, text="Back", command=self.back_to_main, bg="red", fg="white", font=("Arial", 12, "bold"))
        back_button.place(x=175, y=15, width=80, height=30)

        # Main Frame for Admin Interface
        main_frame2 = Frame(background_img_main_position, bd=2, bg="orange")
        main_frame2.place(x=300, y=5, width=400, height=50)

        # Main Title Label
        main_title = Label(main_frame2, text="Admin Interface", bg="orange", fg="white", font=("New Time Roman", 20, "bold"))
        main_title.place(x=5, y=2, width=400, height=40)

        # Display username on the top right corner
        self.username_label = Label(self.root, text=f"Logged in as: {username}", bg="orange", fg="white", font=("Arial", 12))
        self.username_label.place(x=800, y=15)
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------

        self.report_frame = Frame(self.root, bd=2, bg="orange")
        self.report_frame.place(x=450, y=70, width=560, height=510)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------

        # Left Frame for "Make Report"
        make_report_frame = LabelFrame(self.root, text="Report Form", bg="white", fg="black")
        make_report_frame.place(x=25, y=75, width=420, height=485)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------

        # Create a LabelFrame for "Student Information" inside the make_report_frame
        student_info_frame = LabelFrame(make_report_frame, text="Student Information", bg="white", fg="black")
        student_info_frame.place(x=5, y=5, width=405, height=100)  # Adjusted width to fit well inside make_report_frame

        # Label and Entry for Student ID
        student_id_label = Label(student_info_frame, text="Student ID:", bg="white", fg="black")
        student_id_label.place(x=10, y=10)  # Position for the label

        self.student_id_entry = Entry(student_info_frame, width=15)  # Adjusted width of the entry
        self.student_id_entry.place(x=100, y=10)  # Adjusted position to align with the label

        # Create the "Search ID" button
        search_button = Button(student_info_frame, text="Search ID", bg="orange", fg="white", width=10, command=self.search_student_info)  # Set width for uniformity
        search_button.place(x=270, y=10)  # Positioning the button to the right of the entry

        # Label to display the Student Name
        student_name_label = Label(student_info_frame, text="Student Found:", bg="white", fg="black")
        student_name_label.place(x=10, y=40)  # Position for the name label
        self.student_name_display = Label(student_info_frame, text="", bg="white", fg="black")  # Label to display name
        self.student_name_display.place(x=150, y=40)  # Position for the name display

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------
         # Create a LabelFrame for "Student Information" inside the make_report_frame
        date_info_frame = LabelFrame(make_report_frame, text="Month / Year", bg="white", fg="black")
        date_info_frame.place(x=5, y=110, width=405, height=50)

        # Label for Month
        month_label = Label(date_info_frame, text="Select Month:", bg="white", fg="black")
        month_label.place(x=10, y=5)  # Positioning for Month label

        # Dropdown menu for Month using Combobox
        self.month_var = StringVar()
        self.month_var.set("Select Month")  # Default option
        months = [str(i) for i in range(1, 13)]  # List of months (1 to 12)
        month_dropdown = ttk.Combobox(date_info_frame, textvariable=self.month_var, values=months, width=13)
        month_dropdown.place(x=100, y=5)  # Positioning for Month dropdown

        # Label for Year
        year_label = Label(date_info_frame, text="Select Year:", bg="white", fg="black")
        year_label.place(x=220, y=5)  # Positioning for Year label

        # Dropdown menu for Year using Combobox
        self.year_var = StringVar()
        self.year_var.set("Select Year")  # Default option
        years = [str(i) for i in range(2020, 2031)]  # Example range of years (2020 to 2030)
        year_dropdown = ttk.Combobox(date_info_frame, textvariable=self.year_var, values=years,width=13)
        year_dropdown.place(x=300, y=5)  # Positioning for Year dropdown


#------------------------------------------------------------------------------------------------------------------------------------------------------------------------

        # LabelFrame for "Course" inside the make_report_frame, below the student_info_frame
        course_frame = LabelFrame(make_report_frame, text="Course", bg="white", fg="black")
        course_frame.place(x=5, y=160, width=405, height=160)  # Positioned below student_info_frame

        # Label and Combobox for Course Name
        course_name_label = Label(course_frame, text="Course Name:", bg="white", fg="black")
        course_name_label.place(x=10, y=10)

        # Assuming course_frame is defined elsewhere in your code
        self.course_name_combobox = ttk.Combobox(course_frame, values=[], width=37, state="readonly")
        self.course_name_combobox.place(x=150, y=10)
        self.course_name_combobox.set("Select Course")  # Set the default option to "Select Course"
        # Set the default option (index 0)

        # Label for "Course Selected:"
        selected_courses_label = Label(course_frame, text="Course Selected:", bg="white", fg="black")
        selected_courses_label.place(x=10, y=45)  # Position the label to the left of the Listbox

        # Listbox with scrollbar to display selected courses
        self.selected_courses_listbox = Listbox(course_frame, height=3, width=40)

        # Create a Scrollbar widget and attach it to the course_frame
        scrollbar = Scrollbar(course_frame, orient="vertical", command=self.selected_courses_listbox.yview)
        scrollbar.place(x=375, y=45, height=55)  # Adjust the 'x' position based on listbox's width

        # Configure the Listbox to use the scrollbar
        self.selected_courses_listbox.config(yscrollcommand=scrollbar.set)

        # Position the Listbox in the course_frame
        self.selected_courses_listbox.place(x=150, y=45)

        # Button to delete selected course from the Listbox
        self.delete_course_button = Button(course_frame, text="Delete Selected Course", command=self.delete_selected_course)
        self.delete_course_button.place(x=150, y=100)  # 5 pixels gap from the Listbox

        # Bind the Combobox selection event to call add_course function
        self.course_name_combobox.bind("<<ComboboxSelected>>", lambda event: self.add_course())


#------------------------------------------------------------------------------------------------------------------------------------------------------------------------
       # LabelFrame for "Emotion Status" below the course_frame
        emotion_status_frame = LabelFrame(make_report_frame, text="Emotion Status", bg="white", fg="black")
        emotion_status_frame.place(x=5, y=320, width=405, height=50)  # Adjusted height to fit in a single row

        # Variables to hold the state of the checkboxes
        self.detail_var = BooleanVar()
        self.overall_var = BooleanVar()
        self.table_var = BooleanVar()

        # Checkbuttons for "Detail", "Overall", and "Table" in a horizontal order
        self.detail_checkbox = Checkbutton(emotion_status_frame, text="Chart (Detailed)", variable=self.detail_var, bg="white")
        self.detail_checkbox.place(x=20, y=3)

        self.overall_checkbox = Checkbutton(emotion_status_frame, text="Chart (Overall)", variable=self.overall_var, bg="white")
        self.overall_checkbox.place(x=150, y=3)

        self.table_checkbox = Checkbutton(emotion_status_frame, text="Status in Table", variable=self.table_var, bg="white")
        self.table_checkbox.place(x=280, y=3)

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # LabelFrame for "Report Generate" below the emotion_status_frame
        report_generate_frame = LabelFrame(make_report_frame, text="Report Generate", bg="white", fg="black")
        report_generate_frame.place(x=5, y=370, width=405, height=90)  # Positioned below emotion_status_frame

        # Configure columns to expand and fill available space
        report_generate_frame.columnconfigure(0, weight=1)
        report_generate_frame.columnconfigure(1, weight=1)

        # Button for "Display Attendance" (formerly "Display Info")
        display_attendance_button = Button(report_generate_frame, bg="orange", fg="white", text="Display Attendance",command=self.display_student_attendance)
        display_attendance_button.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        # Button for "Display Emotion"
        display_emotion_button = Button(report_generate_frame, bg="orange", fg="white", text="Display Emotion", command=self.display_student_emotion)
        display_emotion_button.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")

        # Button for "Generate Attendance" (formerly "Generate Report")
        generate_attendance_button = Button(report_generate_frame, bg="orange", fg="white", text="Generate Attendance", command=self.generate_report)
        generate_attendance_button.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        # Button for "Generate Emotion"
        generate_emotion_button = Button(report_generate_frame, bg="orange", fg="white", text="Generate Emotion")
        generate_emotion_button.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")



        
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------

      # Right Frame for PDF report preview
        self.preview_frame = Frame(self.report_frame, bg="white", bd=2)  # Set background to white
        self.preview_frame.place(x=5, y=5, width=550, height=490)

        # Create a Canvas for the scrollable area
        self.canvas = Canvas(self.preview_frame, bg="white")  # Set canvas background to white
        self.canvas.pack(side="left", fill="both", expand=True)

        # Create a frame to hold the content inside the Canvas
        self.content_frame = Frame(self.canvas, bg="white")  # Set content frame background to white
        self.canvas.create_window((0, 0), window=self.content_frame, anchor="nw")

        # Create vertical Scrollbar and link it to the Canvas
        self.y_scrollbar = Scrollbar(self.preview_frame, orient="vertical", command=self.canvas.yview)
        self.y_scrollbar.pack(side="right", fill="y")

        # Configure the canvas to use the vertical scrollbar only
        self.canvas.configure(yscrollcommand=self.y_scrollbar.set)  # Updated to link yscrollbar only

        # Bind the content frame size to update the scroll region
        self.content_frame.bind("<Configure>", self.update_scroll_region)




        

        




       

       

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # Function to update the scroll region
    def display_student_attendance(self):
        # Clear previous content in the content_frame
        for widget in self.content_frame.winfo_children():
            widget.destroy()

        # Fetch student ID, name, month, and year
        student_id = self.student_id_entry.get()
        student_name = self.student_name_display.cget("text")
        selected_month = self.month_var.get()  # Fetch selected month
        selected_year = self.year_var.get()  # Fetch selected year

        # Check if student ID is empty
        if not student_id:
            messagebox.showwarning("Input Error", "Please enter a Student ID.")
            return  # Exit the function if ID is empty

        # Add the title label for the report and center it
        Label(self.content_frame, text="Report Form", bg="white", fg="black", font=("Arial", 16, "bold")).pack(padx=10, pady=(10, 5), anchor="center", fill="x")

        # Create a frame for the table-like layout with a fixed width of 500
        table_frame = Frame(self.content_frame, bg="white", width=500)
        table_frame.pack(padx=10, pady=5, anchor="center")  # Set padding and anchor

        # Set the column width to 250 for both columns
        table_frame.columnconfigure(0, minsize=250)  # Set minimum width for column 0
        table_frame.columnconfigure(1, minsize=250)  # Set minimum width for column 1

        # Create rows for each label (Name, ID, Month, Year)
        # Row 1: Name and ID
        Label(table_frame, text="Name: " + student_name, bg="white", fg="black").grid(row=0, column=0, sticky="w")  # Name label
        Label(table_frame, text="ID: " + student_id, bg="white", fg="black").grid(row=0, column=1, sticky="w")  # ID label

        # Row 2: Month and Year
        Label(table_frame, text="Month: " + selected_month, bg="white", fg="black").grid(row=1, column=0, pady=5, sticky="w")  # Month label
        Label(table_frame, text="Year: " + selected_year, bg="white", fg="black").grid(row=1, column=1, pady=5, sticky="w")  # Year label

        # Display selected courses label, centered and with font size 14
        Label(self.content_frame, text="Courses Attendance", bg="white", fg="black", font=("Arial", 14)).pack(anchor="center", padx=10, pady=5)        

        selected_courses = self.selected_courses_listbox.get(0, 'end')

        for course in selected_courses:
            Label(self.content_frame, text="- " + course, bg="white", fg="black").pack(anchor="w", padx=20)
            
            # Fetch the attendance data for each course
            attendance_data = self.fetch_attendance_data(student_id, course, selected_month, selected_year)
            
            # Create a frame for the attendance table
            table_frame = Frame(self.content_frame, bg="white")
            table_frame.pack(padx=10, pady=(5, 10))  # Fill both vertically and horizontally

            # Table headers
            headers = ["Date", "Hour", "Status"]
            header_bg_color = "lightgray"  # Background color for headers

            for row, header in enumerate(headers):
                Label(table_frame, text=header, bg=header_bg_color, fg="black", font=("Arial", 10, "bold")).grid(row=0, column=row, sticky="nsew", padx=5, pady=5)

            # Populate the table with attendance data
            for row_index, (date, attendance_status, course_hour) in enumerate(attendance_data, start=1):
                status_mark = ''
                if attendance_status == "Present":
                    status_mark = "P"
                elif attendance_status == "Absent":
                    status_mark = "A"
                elif attendance_status == "Half-Absent":
                    status_mark = "P/A"

                # Directly add data without rotation
                Label(table_frame, text=date, bg="white", fg="black").grid(row=row_index, column=0, padx=5, pady=5)
                Label(table_frame, text=course_hour, bg="white", fg="black").grid(row=row_index, column=1, padx=5, pady=5)
                Label(table_frame, text=status_mark, bg="white", fg="black").grid(row=row_index, column=2, padx=5, pady=5)

            # Configure grid weights to allow proper expansion
            for i in range(len(attendance_data) + 1):  # Number of rows
                table_frame.grid_rowconfigure(i, weight=1)

            for i in range(len(headers)):  # Number of columns
                table_frame.grid_columnconfigure(i, weight=1)

    
    
    def update_scroll_region(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))  # Update the scroll region to encompass the content_frame

    def fetch_attendance_data_report(self, student_id, course, month, year):
            connection = mysql.connector.connect(
                host='localhost',
                user='root',
                password='Nightcore_1134372019!',
                database='attendnow'
            )
            
            cursor = connection.cursor()
            
            # Modified SQL query to filter by student_id, course, month, and year
            query = """
                SELECT date, course_hour , attendance_status
                FROM attendance_status 
                WHERE student_id = %s 
                AND course = %s 
                AND YEAR(date) = %s 
                AND MONTH(date) = %s
                ORDER BY date
            """
            
            # Execute query with student_id, course, year, and month as parameters
            cursor.execute(query, (student_id, course, year, month))
            
            # Fetch all rows of the result
            attendance_data = cursor.fetchall()
            cursor.close()
            connection.close()
            
            return attendance_data


    def fetch_attendance_data(self, student_id, course, month, year):
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='Nightcore_1134372019!',
            database='attendnow'
        )
        
        cursor = connection.cursor()
        
        # Modified SQL query to filter by student_id, course, month, and year
        query = """
            SELECT date, attendance_status, course_hour 
            FROM attendance_status 
            WHERE student_id = %s 
            AND course = %s 
            AND YEAR(date) = %s 
            AND MONTH(date) = %s
            ORDER BY date
        """
        
        # Execute query with student_id, course, year, and month as parameters
        cursor.execute(query, (student_id, course, year, month))
        
        # Fetch all rows of the result
        attendance_data = cursor.fetchall()
        cursor.close()
        connection.close()
        
        return attendance_data


    def generate_report(self):
        # Fetch student details
        student_id = self.student_id_entry.get()
        student_name = self.student_name_display.cget("text")
        selected_month = self.month_var.get()
        selected_year = self.year_var.get()

        # Convert month number to name
        month_names = ["Invalid month", "January", "February", "March", "April", "May", "June",
                    "July", "August", "September", "October", "November", "December"]
        selected_month_text = month_names[int(selected_month)] if selected_month.isdigit() and 1 <= int(selected_month) <= 12 else "Invalid month"

        # Check for missing Student ID
        if not student_id:
            messagebox.showwarning("Input Error", "Please enter a Student ID.")
            return

        # Define folder name
        folder_name = f"{student_name}_{student_id}_{selected_month_text}_{selected_year}_Emotion"
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        # Fetch the list of selected courses from the listbox
        selected_courses = self.selected_courses_listbox.get(0, 'end')

        for course in selected_courses:
            # Create a new Excel workbook and sheet for each course
            excel_filename = f"{folder_name}/{student_name}_{student_id}_{course}_{selected_month_text}_{selected_year}.xlsx"
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = f"{course} Report"

            # Merge the cells A1 to D1 for the title
            sheet.merge_cells('A1:D1')
            sheet['A1'] = "Report Form Attendance"
            sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Merge A2:B2 for Name and C2:D2 for ID
            sheet.merge_cells('A2:B2')
            sheet['A2'] = f"Name: {student_name}"
            sheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

            sheet.merge_cells('C2:D2')
            sheet['C2'] = f"ID: {student_id}"
            sheet['C2'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

            # Merge A3:B3 for Month and C3:D3 for Year
            sheet.merge_cells('A3:B3')
            sheet['A3'] = f"Month: {selected_month_text}"
            sheet['A3'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

            sheet.merge_cells('C3:D3')
            sheet['C3'] = f"Year: {selected_year}"
            sheet['C3'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

            # Merge A4:D4 for Course name
            sheet.merge_cells('A4:D4')
            sheet['A4'] = f"Course: {course}"
            sheet['A4'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

            # Add a space before the next section
            row_start = 6

            # Merge cells A6:D6 for Attendance Status Title
            sheet.merge_cells('A6:D6')
            sheet['A6'] = "Attendance Status"
            sheet['A6'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            # Set column widths for A, B, C, D to 200 pixels (approximately 200 in Excel units)
            sheet.column_dimensions['A'].width = 28  # Approximation for pixels
            sheet.column_dimensions['B'].width = 28
            sheet.column_dimensions['C'].width = 28
            sheet.column_dimensions['D'].width = 28

            # Fetch attendance data for the course
            attendance_data = self.fetch_attendance_data_report(student_id, course, selected_month, selected_year)

            # Define headers for attendance data
            headers = ["Date", "Hour", "Status"]
            sheet[f"A{row_start + 1}"] = headers[0]
            sheet[f"B{row_start + 1}"] = headers[1]
            sheet[f"C{row_start + 1}"] = headers[2]

            # Populate the attendance data rows
            row = row_start + 2
            for row_data in attendance_data:
                sheet[f"A{row}"] = row_data[0]  # Date
                sheet[f"B{row}"] = row_data[1]  # Hour
                sheet[f"C{row}"] = row_data[2]  # Status
                row += 1

            # Save the Excel file
            wb.save(excel_filename)
            print(f"Excel file saved as {excel_filename}")

    


    def display_student_emotion(self):
        # Clear previous content in the content_frame
        for widget in self.content_frame.winfo_children():
            widget.destroy()

        # Fetch student ID, name, month, and year
        student_id = self.student_id_entry.get()
        student_name = self.student_name_display.cget("text")
        selected_month = self.month_var.get()  # Fetch selected month
        selected_year = self.year_var.get()  # Fetch selected year

        # Check if student ID is empty
        if not student_id:
            messagebox.showwarning("Input Error", "Please enter a Student ID.")
            return  # Exit the function if ID is empty

        # Add the title label for the report and center it
        Label(self.content_frame, text="Report Form", bg="white", fg="black", font=("Arial", 16, "bold")).pack(padx=10, pady=(10, 5), anchor="center", fill="x")

        # Create a frame for the table-like layout with a fixed width of 500
        table_frame = Frame(self.content_frame, bg="white", width=500)
        table_frame.pack(padx=10, pady=5, anchor="center")  # Set padding and anchor

        # Set the column width to 250 for both columns
        table_frame.columnconfigure(0, minsize=250)  # Set minimum width for column 0
        table_frame.columnconfigure(1, minsize=250)  # Set minimum width for column 1

        # Create rows for each label (Name, ID, Month, Year)
        # Row 1: Name and ID
        Label(table_frame, text="Name: " + student_name, bg="white", fg="black").grid(row=0, column=0, sticky="w")  # Name label
        Label(table_frame, text="ID: " + student_id, bg="white", fg="black").grid(row=0, column=1, sticky="w")  # ID label

        # Row 2: Month and Year
        Label(table_frame, text="Month: " + selected_month, bg="white", fg="black").grid(row=1, column=0, pady=5, sticky="w")  # Month label
        Label(table_frame, text="Year: " + selected_year, bg="white", fg="black").grid(row=1, column=1, pady=5, sticky="w")  # Year label

        # Display Emotion Status label, centered and with font size 14
        Label(self.content_frame, text="Emotion Status", bg="white", fg="black", font=("Arial", 14)).pack(anchor="center", padx=10, pady=5)
        
        if self.detail_var.get():
            # Get the list of selected courses from the listbox
            selected_courses = self.selected_courses_listbox.get(0, 'end')

            # Loop through each course and call the function to display the emotion status chart
            for course in selected_courses:
                self.get_emotion_status(student_id, student_name, selected_month, selected_year, course)

        if self.overall_var.get():
            # Get the list of selected courses from the listbox
            selected_courses = self.selected_courses_listbox.get(0, 'end')

            for course in selected_courses:
                self.emotion_status_overall(student_id, student_name, selected_month, selected_year, course)

        if self.table_var.get():
            selected_courses = self.selected_courses_listbox.get(0, 'end')

            for course in selected_courses:
                # Fetch emotional data
                emotional_data = self.emotion_status_table(student_id, student_name, selected_month, selected_year, course)

                # Create a frame for the emotion status table
                table_frame = Frame(self.content_frame, bg="white")
                table_frame.pack(padx=10, pady=(5, 10))  # Add some padding and fill horizontally

                # Add the course name as a heading in row 0, spanning across all columns
                Label(table_frame, text=course, font=("Arial", 12, "bold"), bg="white", fg="black").grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 5))

                # Create table headers in row 1
                Label(table_frame, text="Date", bg="white", fg="black").grid(row=1, column=0, padx=5, pady=5)
                Label(table_frame, text="Neutral", bg="white", fg="black").grid(row=1, column=1, padx=5, pady=5)
                Label(table_frame, text="Happy", bg="white", fg="black").grid(row=1, column=2, padx=5, pady=5)
                Label(table_frame, text="Sad", bg="white", fg="black").grid(row=1, column=3, padx=5, pady=5)
                Label(table_frame, text="Fear", bg="white", fg="black").grid(row=1, column=4, padx=5, pady=5)
                Label(table_frame, text="Surprise", bg="white", fg="black").grid(row=1, column=5, padx=5, pady=5)
                Label(table_frame, text="Angry", bg="white", fg="black").grid(row=1, column=6, padx=5, pady=5)

                # Populate the table with emotional data, starting from row 2
                for row_index, (date, neutral, happy, sad, fear, surprise, angry) in enumerate(emotional_data, start=2):
                    Label(table_frame, text=date, bg="white", fg="black").grid(row=row_index, column=0, padx=5, pady=5)
                    Label(table_frame, text=neutral, bg="white", fg="black").grid(row=row_index, column=1, padx=5, pady=5)
                    Label(table_frame, text=happy, bg="white", fg="black").grid(row=row_index, column=2, padx=5, pady=5)
                    Label(table_frame, text=sad, bg="white", fg="black").grid(row=row_index, column=3, padx=5, pady=5)
                    Label(table_frame, text=fear, bg="white", fg="black").grid(row=row_index, column=4, padx=5, pady=5)
                    Label(table_frame, text=surprise, bg="white", fg="black").grid(row=row_index, column=5, padx=5, pady=5)
                    Label(table_frame, text=angry, bg="white", fg="black").grid(row=row_index, column=6, padx=5, pady=5)

                # Update scroll region to encompass all items
                self.update_scroll_region()




    def emotion_status_table(self, student_id, student_name, selected_month, selected_year, course): 
        # Connect to the MySQL database
        connection = mysql.connector.connect(
            host="localhost",  # Replace with your host if different
            user="root",
            password="Nightcore_1134372019!",
            database="attendnow"
        )

        cursor = connection.cursor()

        # SQL query to extract emotion data by student name, ID, month, and year
        query = """
            SELECT date, neutral, happy, sad, fear, surprise, angry
            FROM student_emotion
            WHERE student_name = %s 
            AND student_id = %s
            AND course = %s 
            AND MONTH(date) = %s
            AND YEAR(date) = %s
            ORDER BY date;
        """

        cursor.execute(query, (student_name, student_id, course,selected_month, selected_year))
        results = cursor.fetchall()
        cursor.close()
        connection.close()

        return results

    def emotion_status_overall(self, student_id, student_name, selected_month, selected_year, course):
        # Connect to the MySQL database
        connection = mysql.connector.connect(
            host="localhost",  # Replace with your host if different
            user="root",
            password="Nightcore_1134372019!",
            database="attendnow"
        )
        
        cursor = connection.cursor()

        # SQL query to extract emotion data by student name, ID, month, and year
        query = """
            SELECT date, neutral, happy, sad, fear, surprise, angry
            FROM student_emotion
            WHERE student_name = %s 
            AND student_id = %s
            AND course = %s 
            AND MONTH(date) = %s
            AND YEAR(date) = %s
            ORDER BY date;
        """
        
        cursor.execute(query, (student_name, student_id, course,selected_month, selected_year))
        results = cursor.fetchall()
        
        # Close the database connection 
        connection.close()

        if not results:
            month_name = calendar.month_name[int(selected_month)]
            messagebox.showinfo("Info", f"No emotion data found for the selected student in {month_name}, {selected_year}.")
            return

        # Extracting data for the chart
        days = []  
        max_emotion_values = []
        max_emotion_labels = []

        # Define emotion color mapping
        emotion_colors = {
            'Neutral': 'green',
            'Happy': 'yellow',
            'Sad': 'blue',
            'Fear': 'purple',
            'Surprise': 'orange',
            'Angry': 'red'
        }

        for row in results:
            # Extract the day from the date
            date_str = str(row[0])  # Assuming the date is in a datetime format
            day = date_str.split('-')[2]  # Get the day (DD) from YYYY-MM-DD
            
            days.append(day)  # Append the day instead of the full date
            
            # Find the maximum emotion value and its corresponding label
            emotion_values = [row[1], row[2], row[3], row[4], row[5], row[6]]
            max_value = max(emotion_values)
            max_index = emotion_values.index(max_value)

            max_emotion_values.append(max_value)
            
            # Map emotion index to labels
            emotions = ['Neutral', 'Happy', 'Sad', 'Fear', 'Surprise', 'Angry']
            max_emotion_labels.append(emotions[max_index])

        # Plotting the bar chart with only the max emotions
        figure, ax = plt.subplots(figsize=(5, 4))  # Set the figure size to fit your needs

        # Assign the color based on the max emotion
        colors = [emotion_colors[label] for label in max_emotion_labels]

        bars = ax.bar(days, max_emotion_values, color=colors)  # Use the corresponding color for each max emotion

        # Setting the labels
        ax.set_xlabel('Day (DD)')
        ax.set_ylabel('Emotion Value')
        ax.set_title(f'Max Emotion Status for {student_name} ({calendar.month_name[int(selected_month)]}, {selected_year}) ')

        # Updated x-ticks and labels
        ax.set_xticks(range(len(days)))
        ax.set_xticklabels(days, rotation=90)  # Rotate labels for better visibility

        # Create a legend for the emotion colors
        handles = [plt.Rectangle((0, 0), 1, 1, color=color) for color in emotion_colors.values()]
        ax.legend(handles, emotion_colors.keys(), title="Emotions")

        # Displaying the chart in the Tkinter Frame (main_frame)
        canvas = FigureCanvasTkAgg(figure, master=self.content_frame)
        canvas.draw()

        # Set the canvas size explicitly
        canvas.get_tk_widget().config(width=500, height=300)  # Set the canvas size

        # Pack the canvas to fill the main_frame
        canvas.get_tk_widget().pack(fill="both", expand=True)


    def get_emotion_status(self, student_id, student_name, selected_month, selected_year, course):
        # Connect to the MySQL database
        connection = mysql.connector.connect(
            host="localhost",  # Replace with your host if different
            user="root",
            password="Nightcore_1134372019!",
            database="attendnow"
        )
        
        cursor = connection.cursor()

        # SQL query to extract emotion data by student name, ID, month, and year
        query = """
            SELECT date, neutral, happy, sad, fear, surprise, angry
            FROM student_emotion
            WHERE student_name = %s 
            AND student_id = %s
            AND course = %s 
            AND MONTH(date) = %s
            AND YEAR(date) = %s
            ORDER BY date;
        """
        
        cursor.execute(query, (student_name, student_id, course,selected_month, selected_year))
        results = cursor.fetchall()
        
        # Close the database connection 
        connection.close()

        if not results:
            messagebox.showinfo("Info", f"No emotion data found for the selected student in {selected_month}, {selected_year}.")
            return

        # Extracting data for the chart
        days = []  # Change this to hold days instead of full dates
        neutral = []
        happy = []
        sad = []
        fear = []
        surprise = []
        angry = []
        
        for row in results:
            # Extract the day from the date
            date_str = str(row[0])  # Assuming the date is in a datetime format
            day = date_str.split('-')[2]  # Get the day (DD) from YYYY-MM-DD
            
            days.append(day)  # Append the day instead of the full date
            neutral.append(row[1])
            happy.append(row[2])
            sad.append(row[3])
            fear.append(row[4])
            surprise.append(row[5])
            angry.append(row[6])

        # Plotting the vertical stacked bar chart with a fixed width
        figure, ax = plt.subplots(figsize=(5, 4))  # Set width to 5 inches (approximately 500 pixels)

        # Stacking the bars correctly
        ax.bar(days, neutral, label='Neutral', color='green')
        ax.bar(days, happy, bottom=neutral, label='Happy', color='yellow')
        ax.bar(days, sad, bottom=[i + j for i, j in zip(neutral, happy)], label='Sad', color='blue')
        ax.bar(days, fear, bottom=[i + j + k for i, j, k in zip(neutral, happy, sad)], label='Fear', color='purple')
        ax.bar(days, surprise, bottom=[i + j + k + l for i, j, k, l in zip(neutral, happy, sad, fear)], label='Surprise', color='orange')
        ax.bar(days, angry, bottom=[i + j + k + l + m for i, j, k, l, m in zip(neutral, happy, sad, fear, surprise)], label='Angry', color='red')

        # Set x-axis ticks and rotate labels to prevent overlap
        ax.set_xticks(range(len(days)))
        ax.set_xticklabels(days, rotation=90)

        # Setting the labels
        ax.set_xlabel('Day (DD)')
        ax.set_ylabel('Emotion')
        ax.set_title(f'Emotion Status ({selected_month}, {selected_year}) {course}')
        ax.legend()

        # Displaying the chart in the Tkinter Frame (content_frame inside preview_frame)
        canvas = FigureCanvasTkAgg(figure, master=self.content_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)




    def update_scroll_region(self, event=None):
        # Update the scroll region of the canvas
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
        
    
    
    def search_student_info(self):
        student_id = self.student_id_entry.get()  # Get the Student ID from the entry

        # Clear previous selections in the Combobox
        self.course_name_combobox.set("")  # Clear the Combobox display
        self.student_name_display.config(text="")  # Clear previous student name display

        # Connect to the MySQL database
        try:
            connection = mysql.connector.connect(
                host='localhost',  # Update with your host
                user='root',       # Update with your MySQL username
                password='Nightcore_1134372019!',  # Update with your MySQL password
                database='attendnow'  # The database name
            )

            cursor = connection.cursor()

            # SQL query to fetch the student name based on the student ID
            query_name = "SELECT student_name FROM students WHERE student_id = %s LIMIT 1"  # Get the student name
            cursor.execute(query_name, (student_id,))
            
            # Fetch the student name result
            student_name_result = cursor.fetchone()

            if student_name_result:  # If a result is found
                student_name = student_name_result[0]  # Get the student name
                self.student_name_display.config(text=student_name)  # Update the label with the name
            else:
                self.student_name_display.config(text="Not Found")  # Update label if no student found

            # SQL query to fetch the courses based on the student ID
            query_courses = "SELECT course FROM students WHERE student_id = %s"  # Get all courses for the student
            cursor.execute(query_courses, (student_id,))
            
            # Fetch all results for courses
            course_results = cursor.fetchall()

            # Create a list to hold course names
            course_names = []

            if course_results:  # If results are found
                for row in course_results:
                    course = row[0]  # Get the course name from the result
                    course_names.append(course)  # Add the course to the list

            # Update the Combobox with the list of courses
            self.course_name_combobox['values'] = course_names  # Set the new list of courses
            if course_names:  # If there are courses, set the first one as default
                self.course_name_combobox.current(0)

            else:
                self.course_name_combobox['values'] = []  # Clear values if no courses found

        except mysql.connector.Error as err:
            # Show an error message if something goes wrong
            messagebox.showerror("Database Error", f"Error: {err}")
        finally:
            # Close the cursor and connection
            cursor.close()
            connection.close()


    def delete_selected_course(self):
        # Get the currently selected item index
        selected_index = self.selected_courses_listbox.curselection()

        if selected_index:  # Check if an item is selected
            self.selected_courses_listbox.delete(selected_index)  # Delete the selected item
        else:
            messagebox.showwarning("Selection Error", "Please select a course to delete.")  # Warn if nothing is selected


    # Define the add_course method within the class
    def add_course(self):
        selected_course = self.course_name_combobox.get()  # Get the selected course from Combobox
        if selected_course not in self.selected_courses_listbox.get(0, END):  # Check if the course is not already added
            self.selected_courses_listbox.insert(END, selected_course)  # Add the selected course to the Listbox

    def back_to_main(self):
        self.root.destroy()  # Close the current window
        new_window = Tk()  # Create a new Tk window for the admit interface
        admit_interface.Admit_Interface(new_window, self.username)


# Example usage
if __name__ == "__main__":
    root = tk.Tk()
    report_generator = Report_Generater(root, "AdminUser")
    root.mainloop()
