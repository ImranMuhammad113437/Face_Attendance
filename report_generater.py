import tkinter as tk
from tkinter import ttk, Listbox
from tkinter import Frame, Canvas, Scrollbar, Label, messagebox, StringVar, BooleanVar, Entry, Listbox
from tkinter import LabelFrame, Label, Entry, Button, StringVar
from tkinter import *
from PIL import Image, ImageTk
import admit_interface
import mysql.connector  
from tkinter import Label, Frame, messagebox  
import mysql.connector
from tkinter import messagebox  
from fpdf import FPDF
from pdf2image import convert_from_path
from reportlab.pdfgen import canvas
from datetime import datetime  
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import calendar
import os
from tkinter import Frame, Label, Tk
from PIL import Image, ImageDraw, ImageFont, ImageTk

from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib.colors import gray
from reportlab.pdfgen import canvas
from reportlab.lib import colors

from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
import os
from datetime import datetime  
from reportlab.pdfgen import canvas
from tkinter import messagebox
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from tkinter import messagebox
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from tkinter import messagebox
import os
import openpyxl
from tkinter import messagebox



class Report_Generater:
    def __init__(self, root, username):
        self.root = root
        self.username = username  
        self.root.geometry("1024x590+0+0")
        self.root.title("AttendNow")

        
        background_img_main = Image.open(r"Image\Background.png")
        background_img_main = background_img_main.resize((1024, 590), Image.Resampling.LANCZOS)
        self.photo_background_img_main = ImageTk.PhotoImage(background_img_main)
        background_img_main_position = Label(self.root, image=self.photo_background_img_main)
        background_img_main_position.place(x=0, y=0, width=1024, height=590)
        

        
        left_title = Image.open(r"Image\LogoTitle_Left Top.png")
        self.photoleft_title = ImageTk.PhotoImage(left_title)
        left_title_position = Label(self.root, image=self.photoleft_title)
        left_title_position.place(x=0, y=0, width=163, height=60)

        
        back_button = Button(self.root, text="Back", command=self.back_to_main, bg="red", fg="white", font=("Arial", 12, "bold"))
        back_button.place(x=175, y=15, width=80, height=30)

        
        main_frame2 = Frame(background_img_main_position, bd=2, bg="orange")
        main_frame2.place(x=300, y=5, width=400, height=50)

        
        main_title = Label(main_frame2, text="Admin Interface", bg="orange", fg="white", font=("New Time Roman", 20, "bold"))
        main_title.place(x=5, y=2, width=400, height=40)

        
        self.username_label = Label(self.root, text=f"Logged in as: {username}", bg="orange", fg="white", font=("Arial", 12))
        self.username_label.place(x=800, y=15)


        self.report_frame = Frame(self.root, bd=2, bg="orange")
        self.report_frame.place(x=450, y=70, width=560, height=510)



        
        make_report_frame = LabelFrame(self.root, text="Report Form", bg="white", fg="black")
        make_report_frame.place(x=25, y=75, width=420, height=485)



        
        student_info_frame = LabelFrame(make_report_frame, text="Student Information", bg="white", fg="black")
        student_info_frame.place(x=5, y=5, width=405, height=100)  

        
        student_id_label = Label(student_info_frame, text="Student ID:", bg="white", fg="black")
        student_id_label.place(x=10, y=10)  

        self.student_id_entry = Entry(student_info_frame, width=15)  
        self.student_id_entry.place(x=100, y=10)  

        
        search_button = Button(student_info_frame, text="Search ID", bg="orange", fg="white", width=10, command=self.search_student_info)  
        search_button.place(x=270, y=10)  

        
        student_name_label = Label(student_info_frame, text="Student Found:", bg="white", fg="black")
        student_name_label.place(x=10, y=40)  
        self.student_name_display = Label(student_info_frame, text="", bg="white", fg="black")  
        self.student_name_display.place(x=150, y=40)  


         
        date_info_frame = LabelFrame(make_report_frame, text="Month / Year", bg="white", fg="black")
        date_info_frame.place(x=5, y=110, width=405, height=50)

        
        month_label = Label(date_info_frame, text="Select Month:", bg="white", fg="black")
        month_label.place(x=10, y=5)  

        
        self.month_var = StringVar()
        self.month_var.set("Select Month")  
        months = [str(i) for i in range(1, 13)]  
        month_dropdown = ttk.Combobox(date_info_frame, textvariable=self.month_var, values=months, width=13)
        month_dropdown.place(x=100, y=5)  

        
        year_label = Label(date_info_frame, text="Select Year:", bg="white", fg="black")
        year_label.place(x=220, y=5)  

        
        self.year_var = StringVar()
        self.year_var.set("Select Year")  
        years = [str(i) for i in range(2020, 2031)]  
        year_dropdown = ttk.Combobox(date_info_frame, textvariable=self.year_var, values=years,width=13)
        year_dropdown.place(x=300, y=5)  




        
        course_frame = LabelFrame(make_report_frame, text="Course", bg="white", fg="black")
        course_frame.place(x=5, y=160, width=405, height=160)  

        
        course_name_label = Label(course_frame, text="Course Name:", bg="white", fg="black")
        course_name_label.place(x=10, y=10)

        
        self.course_name_combobox = ttk.Combobox(course_frame, values=[], width=37, state="readonly")
        self.course_name_combobox.place(x=150, y=10)
        self.course_name_combobox.set("Select Course")  
        

        
        selected_courses_label = Label(course_frame, text="Course Selected:", bg="white", fg="black")
        selected_courses_label.place(x=10, y=45)  

        
        self.selected_courses_listbox = Listbox(course_frame, height=3, width=40)

        
        scrollbar = Scrollbar(course_frame, orient="vertical", command=self.selected_courses_listbox.yview)
        scrollbar.place(x=375, y=45, height=55)  

        
        self.selected_courses_listbox.config(yscrollcommand=scrollbar.set)

        
        self.selected_courses_listbox.place(x=150, y=45)

        
        self.delete_course_button = Button(course_frame, text="Delete Selected Course", command=self.delete_selected_course)
        self.delete_course_button.place(x=150, y=100)  

        
        self.course_name_combobox.bind("<<ComboboxSelected>>", lambda event: self.add_course())



       
        emotion_status_frame = LabelFrame(make_report_frame, text="Emotion Status", bg="white", fg="black")
        emotion_status_frame.place(x=5, y=320, width=405, height=50)  

        
        self.detail_var = BooleanVar()
        self.overall_var = BooleanVar()
        self.table_var = BooleanVar()

        
        self.detail_checkbox = Checkbutton(emotion_status_frame, text="Chart (Detailed)", variable=self.detail_var, bg="white")
        self.detail_checkbox.place(x=20, y=3)

        self.overall_checkbox = Checkbutton(emotion_status_frame, text="Chart (Overall)", variable=self.overall_var, bg="white")
        self.overall_checkbox.place(x=150, y=3)

        self.table_checkbox = Checkbutton(emotion_status_frame, text="Status in Table", variable=self.table_var, bg="white")
        self.table_checkbox.place(x=280, y=3)


        
        report_generate_frame = LabelFrame(make_report_frame, text="Report Generate", bg="white", fg="black")
        report_generate_frame.place(x=5, y=370, width=405, height=90)  

        
        report_generate_frame.columnconfigure(0, weight=1)
        report_generate_frame.columnconfigure(1, weight=1)

        
        display_attendance_button = Button(report_generate_frame, bg="orange", fg="white", text="Display Attendance",command=self.display_student_attendance)
        display_attendance_button.grid(row=0, column=0, padx=5, pady=5, sticky="nsew")

        
        display_emotion_button = Button(report_generate_frame, bg="orange", fg="white", text="Display Emotion", command=self.display_student_emotion)
        display_emotion_button.grid(row=0, column=1, padx=5, pady=5, sticky="nsew")

        
        generate_attendance_button = Button(report_generate_frame, bg="orange", fg="white", text="Generate Attendance", command=self.generate_report)
        generate_attendance_button.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

        
        generate_emotion_button = Button(report_generate_frame, bg="orange", fg="white", text="Generate Emotion", command=self.generate_emotion_report)
        generate_emotion_button.grid(row=1, column=1, padx=5, pady=5, sticky="nsew")



        


      
        self.preview_frame = Frame(self.report_frame, bg="white", bd=2)  
        self.preview_frame.place(x=5, y=5, width=550, height=490)

        
        self.canvas = Canvas(self.preview_frame, bg="white")  
        self.canvas.pack(side="left", fill="both", expand=True)

        
        self.content_frame = Frame(self.canvas, bg="white")  
        self.canvas.create_window((0, 0), window=self.content_frame, anchor="nw")

        
        self.y_scrollbar = Scrollbar(self.preview_frame, orient="vertical", command=self.canvas.yview)
        self.y_scrollbar.pack(side="right", fill="y")

        
        self.canvas.configure(yscrollcommand=self.y_scrollbar.set)  

        
        self.content_frame.bind("<Configure>", self.update_scroll_region)




        

        




       

       


    def generate_emotion_report(self):
        
        student_id = self.student_id_entry.get()
        student_name = self.student_name_display.cget("text")
        selected_month = self.month_var.get()  
        selected_year = self.year_var.get()  

        
        if not student_id:
            messagebox.showwarning("Input Error", "Please enter a Student ID.")
            return  

        
        selected_courses = self.selected_courses_listbox.get(0, 'end')

        for course in selected_courses:
            
            filename = f"{student_name}_{student_id}_{course}_report.pdf"

            
            pdf = canvas.Canvas(filename, pagesize=A4)
            pdf.setTitle(f"Emotion Report for {course}")
            width, height = A4
            width, height_2 = A4

            
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawCentredString(width / 2, height - 50, "Emotion Report")
            pdf.setFont("Helvetica", 12)
            pdf.drawString(100, height - 80, f"Name: {student_name}")
            pdf.drawString(400, height - 80, f"ID: {student_id}")
            pdf.drawString(100, height - 100, f"Month: {selected_month}")
            pdf.drawString(400, height - 100, f"Year: {selected_year}")
            pdf.drawString(100, height - 120, f"Course: {course}")

            
            if self.detail_var.get():
                pdf.setFont("Helvetica-Bold", 14)
                pdf.drawString(100, height - 160, "Detailed Emotion Status:")
                height -= 180  

                
                emotion_img = self.get_emotion_status_report(student_id, student_name, selected_month, selected_year, course)

                
                if emotion_img:
                    image_path = "emotion_status.jpg"
                    pdf.drawImage(image_path, 100, height - 290, width=400, height=300)  
                    height -= 300  

            
            if self.overall_var.get():
                if self.detail_var.get():
                    pdf.setFont("Helvetica-Bold", 14)
                    pdf.drawString(100, height, "Overall Emotion Status:")
                    height -= 20

                    
                    overall_img = self.emotion_status_overall_report(student_id, student_name, selected_month, selected_year, course)

                    
                    if overall_img:
                        overall_image_path = "emotion_status_overall.jpg"
                        pdf.drawImage(overall_image_path, 100, height - 290, width=400, height=300)  
                        height -= 300  
                else:
                    pdf.setFont("Helvetica-Bold", 14)
                    pdf.drawString(100, height - 160, "Overall Emotion Status:")
                    height -= 180

                    
                    overall_img = self.emotion_status_overall_report(student_id, student_name, selected_month, selected_year, course)

                    
                    if overall_img:
                        overall_image_path = "emotion_status_overall.jpg"
                        pdf.drawImage(overall_image_path, 100, height - 290, width=400, height=300)  
                        height -= 300  

            
            if self.table_var.get():
                if self.detail_var.get() and self.overall_var.get():
                    pdf.showPage()  
                    pdf.setFont("Helvetica-Bold", 14)
                    pdf.drawString(100, height_2 - 80, "Emotion Status Table (Detailed Only):")
                    height_2 -= 20

                    
                    pdf.setFont("Helvetica", 10)
                    column_width = 60  
                    pdf.drawString(100 , height_2 - 100, "Date")
                    pdf.drawString(100 + column_width, height_2 - 100, "Neutral")
                    pdf.drawString(100 + 2 * column_width, height_2 - 100, "Happy")
                    pdf.drawString(100 + 3 * column_width, height_2 - 100, "Sad")
                    pdf.drawString(100 + 4 * column_width, height_2 - 100, "Fear")
                    pdf.drawString(100 + 5 * column_width, height_2 - 100, "Surprise")
                    pdf.drawString(100 + 6 * column_width, height_2 - 100, "Angry")
                    height_2 -= 10  

                    
                    emotion_data = self.emotion_status_table(student_id, student_name, selected_month, selected_year, course)

                    for row in emotion_data:
                        pdf.drawString(100, height_2 - 120, row[0])  
                        pdf.drawString(100 + column_width, height_2 - 120, str(row[1]))  
                        pdf.drawString(100 + 2 * column_width, height_2 - 120, str(row[2]))  
                        pdf.drawString(100 + 3 * column_width, height_2 - 120, str(row[3]))  
                        pdf.drawString(100 + 4 * column_width, height_2 - 120, str(row[4]))  
                        pdf.drawString(100 + 5 * column_width, height_2 - 120, str(row[5]))  
                        pdf.drawString(100 + 6 * column_width, height_2 - 120, str(row[6]))  
                        height_2 -= 20  

                elif (self.detail_var.get() and not self.overall_var.get()) or (not self.detail_var.get() and self.overall_var.get()):
                    pdf.setFont("Helvetica-Bold", 14)
                    pdf.drawString(100, height - 10, "Emotion Status Table (Detailed Only):")
                    height -= 20

                    
                    emotion_data = self.emotion_status_table(student_id, student_name, selected_month, selected_year, course)

                    
                    pdf.setFont("Helvetica-Bold", 10)
                    column_width = 60  
                    pdf.drawString(100, height - 10, "Date")
                    pdf.drawString(100 + column_width, height - 10, "Neutral")
                    pdf.drawString(100 + column_width * 2, height - 10, "Happy")
                    pdf.drawString(100 + column_width * 3, height - 10, "Sad")
                    pdf.drawString(100 + column_width * 4, height - 10, "Fear")
                    pdf.drawString(100 + column_width * 5, height - 10, "Surprise")
                    pdf.drawString(100 + column_width * 6, height - 10, "Angry")
                    height -= 20

                    
                    pdf.setFont("Helvetica", 10)
                    for row in emotion_data:
                        pdf.drawString(100, height - 10, row[0])  
                        pdf.drawString(100 + column_width, height - 10, str(row[1]))  
                        pdf.drawString(100 + column_width * 2, height - 10, str(row[2]))  
                        pdf.drawString(100 + column_width * 3, height - 10, str(row[3]))  
                        pdf.drawString(100 + column_width * 4, height - 10, str(row[4]))  
                        pdf.drawString(100 + column_width * 5, height - 10, str(row[5]))  
                        pdf.drawString(100 + column_width * 6, height - 10, str(row[6]))  
                        height -= 20

                

                elif not self.detail_var.get() and not self.overall_var.get():
                    pdf.setFont("Helvetica-Bold", 14)
                    pdf.drawString(100, height - 160, "Emotion Status Table (No Details or Overall):")
                    height -= 180
                    
                    
                    emotion_data = self.emotion_status_table(student_id, student_name, selected_month, selected_year, course)

                    
                    pdf.setFont("Helvetica-Bold", 10)
                    column_width = 60  
                    pdf.drawString(100, height - 10, "Date")
                    pdf.drawString(100 + column_width, height - 10, "Neutral")
                    pdf.drawString(100 + column_width * 2, height - 10, "Happy")
                    pdf.drawString(100 + column_width * 3, height - 10, "Sad")
                    pdf.drawString(100 + column_width * 4, height - 10, "Fear")
                    pdf.drawString(100 + column_width * 5, height - 10, "Surprise")
                    pdf.drawString(100 + column_width * 6, height - 10, "Angry")
                    height -= 20

                    
                    pdf.setFont("Helvetica", 10)
                    for row in emotion_data:
                        pdf.drawString(100, height - 20, row[0])  
                        pdf.drawString(100 + column_width, height - 20, str(row[1]))  
                        pdf.drawString(100 + column_width * 2, height - 20, str(row[2]))  
                        pdf.drawString(100 + column_width * 3, height - 20, str(row[3]))  
                        pdf.drawString(100 + column_width * 4, height - 20, str(row[4]))  
                        pdf.drawString(100 + column_width * 5, height - 20, str(row[5]))  
                        pdf.drawString(100 + column_width * 6, height - 20, str(row[6]))  
                        height -= 20

            
            pdf.save()

            messagebox.showinfo("Success", "PDF reports generated successfully.")

        
    
    def display_student_attendance(self):
        
        for widget in self.content_frame.winfo_children():
            widget.destroy()

        
        student_id = self.student_id_entry.get()
        student_name = self.student_name_display.cget("text")
        selected_month = self.month_var.get()  
        selected_year = self.year_var.get()  

        
        if not student_id:
            messagebox.showwarning("Input Error", "Please enter a Student ID.")
            return  

        
        Label(self.content_frame, text="Report Form", bg="white", fg="black", font=("Arial", 16, "bold")).pack(padx=10, pady=(10, 5), anchor="center", fill="x")

        
        table_frame = Frame(self.content_frame, bg="white", width=500)
        table_frame.pack(padx=10, pady=5, anchor="center")  

        
        table_frame.columnconfigure(0, minsize=250)  
        table_frame.columnconfigure(1, minsize=250)  

        
        
        Label(table_frame, text="Name: " + student_name, bg="white", fg="black").grid(row=0, column=0, sticky="w")  
        Label(table_frame, text="ID: " + student_id, bg="white", fg="black").grid(row=0, column=1, sticky="w")  

        
        Label(table_frame, text="Month: " + selected_month, bg="white", fg="black").grid(row=1, column=0, pady=5, sticky="w")  
        Label(table_frame, text="Year: " + selected_year, bg="white", fg="black").grid(row=1, column=1, pady=5, sticky="w")  

        
        Label(self.content_frame, text="Courses Attendance", bg="white", fg="black", font=("Arial", 14)).pack(anchor="center", padx=10, pady=5)        

        selected_courses = self.selected_courses_listbox.get(0, 'end')

        for course in selected_courses:
            Label(self.content_frame, text="- " + course, bg="white", fg="black").pack(anchor="w", padx=20)
            
            
            attendance_data = self.fetch_attendance_data(student_id, course, selected_month, selected_year)
            
            
            table_frame = Frame(self.content_frame, bg="white")
            table_frame.pack(padx=10, pady=(5, 10))  

            
            headers = ["Date", "Hour", "Status"]
            header_bg_color = "lightgray"  

            for row, header in enumerate(headers):
                Label(table_frame, text=header, bg=header_bg_color, fg="black", font=("Arial", 10, "bold")).grid(row=0, column=row, sticky="nsew", padx=5, pady=5)

            
            for row_index, (date, attendance_status, course_hour) in enumerate(attendance_data, start=1):
                status_mark = ''
                if attendance_status == "Present":
                    status_mark = "P"
                elif attendance_status == "Absent":
                    status_mark = "A"
                elif attendance_status == "Half-Absent":
                    status_mark = "P/A"

                
                Label(table_frame, text=date, bg="white", fg="black").grid(row=row_index, column=0, padx=5, pady=5)
                Label(table_frame, text=course_hour, bg="white", fg="black").grid(row=row_index, column=1, padx=5, pady=5)
                Label(table_frame, text=status_mark, bg="white", fg="black").grid(row=row_index, column=2, padx=5, pady=5)

            
            for i in range(len(attendance_data) + 1):  
                table_frame.grid_rowconfigure(i, weight=1)

            for i in range(len(headers)):  
                table_frame.grid_columnconfigure(i, weight=1)

    
    
    def update_scroll_region(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))  

    def fetch_attendance_data_report(self, student_id, course, month, year):
            connection = mysql.connector.connect(
                host='localhost',
                user='root',
                password='Nightcore_1134372019!',
                database='attendnow'
            )
            
            cursor = connection.cursor()
            
            
            query = 
            
            
            cursor.execute(query, (student_id, course, year, month))
            
            
            attendance_data = cursor.fetchall()
            cursor.close()
            connection.close()
            
            return attendance_data


    def fetch_attendance_data(self, student_id, course, month, year):
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='Nightcore_1134372019!',
            database='attendnow'
        )
        
        cursor = connection.cursor()
        
        
        query = 
        
        
        cursor.execute(query, (student_id, course, year, month))
        
        
        attendance_data = cursor.fetchall()
        cursor.close()
        connection.close()
        
        return attendance_data


    def generate_report(self):
        
        student_id = self.student_id_entry.get()
        student_name = self.student_name_display.cget("text")
        selected_month = self.month_var.get()
        selected_year = self.year_var.get()

        
        month_names = ["Invalid month", "January", "February", "March", "April", "May", "June",
                    "July", "August", "September", "October", "November", "December"]
        selected_month_text = month_names[int(selected_month)] if selected_month.isdigit() and 1 <= int(selected_month) <= 12 else "Invalid month"

        
        if not student_id:
            messagebox.showwarning("Input Error", "Please enter a Student ID.")
            return

        
        folder_name = f"{student_name}_{student_id}_{selected_month_text}_{selected_year}_Attendance"
        if not os.path.exists(folder_name):
            os.makedirs(folder_name)

        
        selected_courses = self.selected_courses_listbox.get(0, 'end')

        for course in selected_courses:
            
            excel_filename = f"{folder_name}/{student_name}_{student_id}_{course}_{selected_month_text}_{selected_year}.xlsx"
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = f"{course} Report"

            
            sheet.merge_cells('A1:D1')
            sheet['A1'] = "Report Form Attendance"
            sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            
            sheet.merge_cells('A2:B2')
            sheet['A2'] = f"Name: {student_name}"
            sheet['A2'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

            sheet.merge_cells('C2:D2')
            sheet['C2'] = f"ID: {student_id}"
            sheet['C2'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

            
            sheet.merge_cells('A3:B3')
            sheet['A3'] = f"Month: {selected_month_text}"
            sheet['A3'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

            sheet.merge_cells('C3:D3')
            sheet['C3'] = f"Year: {selected_year}"
            sheet['C3'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

            
            sheet.merge_cells('A4:D4')
            sheet['A4'] = f"Course: {course}"
            sheet['A4'].alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')

            
            row_start = 6

            
            sheet.merge_cells('A6:D6')
            sheet['A6'] = "Attendance Status"
            sheet['A6'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

            
            sheet.column_dimensions['A'].width = 28  
            sheet.column_dimensions['B'].width = 28
            sheet.column_dimensions['C'].width = 28
            sheet.column_dimensions['D'].width = 28

            
            attendance_data = self.fetch_attendance_data_report(student_id, course, selected_month, selected_year)

            
            headers = ["Date", "Hour", "Status"]
            sheet[f"A{row_start + 1}"] = headers[0]
            sheet[f"B{row_start + 1}"] = headers[1]
            sheet[f"C{row_start + 1}"] = headers[2]

            
            row = row_start + 2
            for row_data in attendance_data:
                sheet[f"A{row}"] = row_data[0]  
                sheet[f"B{row}"] = row_data[1]  
                sheet[f"C{row}"] = row_data[2]  
                row += 1

            
            wb.save(excel_filename)
            print(f"Excel file saved as {excel_filename}")

    


    def display_student_emotion(self):
        
        for widget in self.content_frame.winfo_children():
            widget.destroy()

        
        student_id = self.student_id_entry.get()
        student_name = self.student_name_display.cget("text")
        selected_month = self.month_var.get()  
        selected_year = self.year_var.get()  

        
        if not student_id:
            messagebox.showwarning("Input Error", "Please enter a Student ID.")
            return  

        
        Label(self.content_frame, text="Report Form", bg="white", fg="black", font=("Arial", 16, "bold")).pack(padx=10, pady=(10, 5), anchor="center", fill="x")

        
        table_frame = Frame(self.content_frame, bg="white", width=500)
        table_frame.pack(padx=10, pady=5, anchor="center")  

        
        table_frame.columnconfigure(0, minsize=250)  
        table_frame.columnconfigure(1, minsize=250)  

        
        
        Label(table_frame, text="Name: " + student_name, bg="white", fg="black").grid(row=0, column=0, sticky="w")  
        Label(table_frame, text="ID: " + student_id, bg="white", fg="black").grid(row=0, column=1, sticky="w")  

        
        Label(table_frame, text="Month: " + selected_month, bg="white", fg="black").grid(row=1, column=0, pady=5, sticky="w")  
        Label(table_frame, text="Year: " + selected_year, bg="white", fg="black").grid(row=1, column=1, pady=5, sticky="w")  

        
        Label(self.content_frame, text="Emotion Status", bg="white", fg="black", font=("Arial", 14)).pack(anchor="center", padx=10, pady=5)
        
        if self.detail_var.get():
            
            selected_courses = self.selected_courses_listbox.get(0, 'end')

            
            for course in selected_courses:
                self.get_emotion_status(student_id, student_name, selected_month, selected_year, course)

        if self.overall_var.get():
            
            selected_courses = self.selected_courses_listbox.get(0, 'end')

            for course in selected_courses:
                self.emotion_status_overall(student_id, student_name, selected_month, selected_year, course)

        if self.table_var.get():
            selected_courses = self.selected_courses_listbox.get(0, 'end')

            for course in selected_courses:
                
                emotional_data = self.emotion_status_table(student_id, student_name, selected_month, selected_year, course)

                
                table_frame = Frame(self.content_frame, bg="white")
                table_frame.pack(padx=10, pady=(5, 10))  

                
                Label(table_frame, text=course, font=("Arial", 12, "bold"), bg="white", fg="black").grid(row=0, column=0, columnspan=6, sticky="w", pady=(0, 5))

                
                Label(table_frame, text="Date", bg="white", fg="black").grid(row=1, column=0, padx=5, pady=5)
                Label(table_frame, text="Neutral", bg="white", fg="black").grid(row=1, column=1, padx=5, pady=5)
                Label(table_frame, text="Happy", bg="white", fg="black").grid(row=1, column=2, padx=5, pady=5)
                Label(table_frame, text="Sad", bg="white", fg="black").grid(row=1, column=3, padx=5, pady=5)
                Label(table_frame, text="Fear", bg="white", fg="black").grid(row=1, column=4, padx=5, pady=5)
                Label(table_frame, text="Surprise", bg="white", fg="black").grid(row=1, column=5, padx=5, pady=5)
                Label(table_frame, text="Angry", bg="white", fg="black").grid(row=1, column=6, padx=5, pady=5)

                
                for row_index, (date, neutral, happy, sad, fear, surprise, angry) in enumerate(emotional_data, start=2):
                    Label(table_frame, text=date, bg="white", fg="black").grid(row=row_index, column=0, padx=5, pady=5)
                    Label(table_frame, text=neutral, bg="white", fg="black").grid(row=row_index, column=1, padx=5, pady=5)
                    Label(table_frame, text=happy, bg="white", fg="black").grid(row=row_index, column=2, padx=5, pady=5)
                    Label(table_frame, text=sad, bg="white", fg="black").grid(row=row_index, column=3, padx=5, pady=5)
                    Label(table_frame, text=fear, bg="white", fg="black").grid(row=row_index, column=4, padx=5, pady=5)
                    Label(table_frame, text=surprise, bg="white", fg="black").grid(row=row_index, column=5, padx=5, pady=5)
                    Label(table_frame, text=angry, bg="white", fg="black").grid(row=row_index, column=6, padx=5, pady=5)

                
                self.update_scroll_region()




    def emotion_status_table(self, student_id, student_name, selected_month, selected_year, course): 
        
        connection = mysql.connector.connect(
            host="localhost",  
            user="root",
            password="Nightcore_1134372019!",
            database="attendnow"
        )

        cursor = connection.cursor()

        
        query = 

        cursor.execute(query, (student_name, student_id, course,selected_month, selected_year))
        results = cursor.fetchall()
        cursor.close()
        connection.close()

        return results

    def emotion_status_overall_report(self, student_id, student_name, selected_month, selected_year, course):
        
        connection = mysql.connector.connect(
            host="localhost",
            user="root",
            password="Nightcore_1134372019!",
            database="attendnow"
        )

        cursor = connection.cursor()

        
        query = 

        cursor.execute(query, (student_name, student_id, course, selected_month, selected_year))
        results = cursor.fetchall()

        
        connection.close()

        if not results:
            month_name = calendar.month_name[int(selected_month)]
            messagebox.showinfo("Info", f"No emotion data found for the selected student in {month_name}, {selected_year}.")
            return

        
        days = []
        max_emotion_values = []
        max_emotion_labels = []

        
        emotion_colors = {
            'Neutral': 'green',
            'Happy': 'yellow',
            'Sad': 'blue',
            'Fear': 'purple',
            'Surprise': 'orange',
            'Angry': 'red'
        }

        for row in results:
            
            date_str = str(row[0])  
            day = date_str.split('-')[2]  

            days.append(day)  

            
            emotion_values = [row[1], row[2], row[3], row[4], row[5], row[6]]
            max_value = max(emotion_values)
            max_index = emotion_values.index(max_value)

            max_emotion_values.append(max_value)

            
            emotions = ['Neutral', 'Happy', 'Sad', 'Fear', 'Surprise', 'Angry']
            max_emotion_labels.append(emotions[max_index])

        
        figure, ax = plt.subplots(figsize=(5, 4))  

        
        colors = [emotion_colors[label] for label in max_emotion_labels]

        ax.bar(days, max_emotion_values, color=colors)  

        
        ax.set_xlabel('Day (DD)')
        ax.set_ylabel('Emotion Value')
        ax.set_title(f'Max Emotion Status for {student_name} ({calendar.month_name[int(selected_month)]}, {selected_year}) ')

        
        ax.set_xticks(range(len(days)))
        ax.set_xticklabels(days, rotation=90)  

        
        handles = [plt.Rectangle((0, 0), 1, 1, color=color) for color in emotion_colors.values()]
        ax.legend(handles, emotion_colors.keys(), title="Emotions")

        
        image_path = "emotion_status_overall.jpg"
        figure.savefig(image_path, format="jpg", dpi=300)

        
        plt.close(figure)

        
        img = Image.open(image_path)
        return img

    
    
    def emotion_status_overall(self, student_id, student_name, selected_month, selected_year, course):
        
        connection = mysql.connector.connect(
            host="localhost",  
            user="root",
            password="Nightcore_1134372019!",
            database="attendnow"
        )
        
        cursor = connection.cursor()

        
        query = 
        
        cursor.execute(query, (student_name, student_id, course,selected_month, selected_year))
        results = cursor.fetchall()
        
        
        connection.close()

        if not results:
            month_name = calendar.month_name[int(selected_month)]
            messagebox.showinfo("Info", f"No emotion data found for the selected student in {month_name}, {selected_year}.")
            return

        
        days = []  
        max_emotion_values = []
        max_emotion_labels = []

        
        emotion_colors = {
            'Neutral': 'green',
            'Happy': 'yellow',
            'Sad': 'blue',
            'Fear': 'purple',
            'Surprise': 'orange',
            'Angry': 'red'
        }

        for row in results:
            
            date_str = str(row[0])  
            day = date_str.split('-')[2]  
            
            days.append(day)  
            
            
            emotion_values = [row[1], row[2], row[3], row[4], row[5], row[6]]
            max_value = max(emotion_values)
            max_index = emotion_values.index(max_value)

            max_emotion_values.append(max_value)
            
            
            emotions = ['Neutral', 'Happy', 'Sad', 'Fear', 'Surprise', 'Angry']
            max_emotion_labels.append(emotions[max_index])

        
        figure, ax = plt.subplots(figsize=(5, 4))  

        
        colors = [emotion_colors[label] for label in max_emotion_labels]

        bars = ax.bar(days, max_emotion_values, color=colors)  

        
        ax.set_xlabel('Day (DD)')
        ax.set_ylabel('Emotion Value')
        ax.set_title(f'Max Emotion Status for {student_name} ({calendar.month_name[int(selected_month)]}, {selected_year}) ')

        
        ax.set_xticks(range(len(days)))
        ax.set_xticklabels(days, rotation=90)  

        
        handles = [plt.Rectangle((0, 0), 1, 1, color=color) for color in emotion_colors.values()]
        ax.legend(handles, emotion_colors.keys(), title="Emotions")

        
        canvas = FigureCanvasTkAgg(figure, master=self.content_frame)
        canvas.draw()

        
        canvas.get_tk_widget().config(width=500, height=300)  

        
        canvas.get_tk_widget().pack(fill="both", expand=True)


    def get_emotion_status_report(self, student_id, student_name, selected_month, selected_year, course):
        
        connection = mysql.connector.connect(
            host="localhost",  
            user="root",
            password="Nightcore_1134372019!",
            database="attendnow"
        )
        
        cursor = connection.cursor()

        
        query = 
        
        cursor.execute(query, (student_name, student_id, course, selected_month, selected_year))
        results = cursor.fetchall()
        
        
        connection.close()

        if not results:
            messagebox.showinfo("Info", f"No emotion data found for the selected student in {selected_month}, {selected_year}.")
            return

        
        days = []  
        neutral = []
        happy = []
        sad = []
        fear = []
        surprise = []
        angry = []
        
        for row in results:
            
            date_str = str(row[0])  
            day = date_str.split('-')[2]  
            
            days.append(day)  
            neutral.append(row[1])
            happy.append(row[2])
            sad.append(row[3])
            fear.append(row[4])
            surprise.append(row[5])
            angry.append(row[6])

        
        figure, ax = plt.subplots(figsize=(5, 4))  

        
        ax.bar(days, neutral, label='Neutral', color='green')
        ax.bar(days, happy, bottom=neutral, label='Happy', color='yellow')
        ax.bar(days, sad, bottom=[i + j for i, j in zip(neutral, happy)], label='Sad', color='blue')
        ax.bar(days, fear, bottom=[i + j + k for i, j, k in zip(neutral, happy, sad)], label='Fear', color='purple')
        ax.bar(days, surprise, bottom=[i + j + k + l for i, j, k, l in zip(neutral, happy, sad, fear)], label='Surprise', color='orange')
        ax.bar(days, angry, bottom=[i + j + k + l + m for i, j, k, l, m in zip(neutral, happy, sad, fear, surprise)], label='Angry', color='red')

        
        ax.set_xticks(range(len(days)))
        ax.set_xticklabels(days, rotation=90)

        
        ax.set_xlabel('Day (DD)')
        ax.set_ylabel('Emotion')
        ax.set_title(f'Emotion Status ({selected_month}, {selected_year}) {course}')
        ax.legend()

        
        image_path = "emotion_status.jpg"
        figure.savefig(image_path, format="jpg", dpi=300)

        
        plt.close(figure)

        
        img = Image.open(image_path)
        return img

    
    def get_emotion_status(self, student_id, student_name, selected_month, selected_year, course):
        
        connection = mysql.connector.connect(
            host="localhost",  
            user="root",
            password="Nightcore_1134372019!",
            database="attendnow"
        )
        
        cursor = connection.cursor()

        
        query = 
        
        cursor.execute(query, (student_name, student_id, course,selected_month, selected_year))
        results = cursor.fetchall()
        
        
        connection.close()

        if not results:
            messagebox.showinfo("Info", f"No emotion data found for the selected student in {selected_month}, {selected_year}.")
            return

        
        days = []  
        neutral = []
        happy = []
        sad = []
        fear = []
        surprise = []
        angry = []
        
        for row in results:
            
            date_str = str(row[0])  
            day = date_str.split('-')[2]  
            
            days.append(day)  
            neutral.append(row[1])
            happy.append(row[2])
            sad.append(row[3])
            fear.append(row[4])
            surprise.append(row[5])
            angry.append(row[6])

        
        figure, ax = plt.subplots(figsize=(5, 4))  

        
        ax.bar(days, neutral, label='Neutral', color='green')
        ax.bar(days, happy, bottom=neutral, label='Happy', color='yellow')
        ax.bar(days, sad, bottom=[i + j for i, j in zip(neutral, happy)], label='Sad', color='blue')
        ax.bar(days, fear, bottom=[i + j + k for i, j, k in zip(neutral, happy, sad)], label='Fear', color='purple')
        ax.bar(days, surprise, bottom=[i + j + k + l for i, j, k, l in zip(neutral, happy, sad, fear)], label='Surprise', color='orange')
        ax.bar(days, angry, bottom=[i + j + k + l + m for i, j, k, l, m in zip(neutral, happy, sad, fear, surprise)], label='Angry', color='red')

        
        ax.set_xticks(range(len(days)))
        ax.set_xticklabels(days, rotation=90)

        
        ax.set_xlabel('Day (DD)')
        ax.set_ylabel('Emotion')
        ax.set_title(f'Emotion Status ({selected_month}, {selected_year}) {course}')
        ax.legend()

        
        canvas = FigureCanvasTkAgg(figure, master=self.content_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)




    def update_scroll_region(self, event=None):
        
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
        
    
    
    def search_student_info(self):
        student_id = self.student_id_entry.get()  

        
        self.course_name_combobox.set("")  
        self.student_name_display.config(text="")  

        
        try:
            connection = mysql.connector.connect(
                host='localhost',  
                user='root',       
                password='Nightcore_1134372019!',  
                database='attendnow'  
            )

            cursor = connection.cursor()

            
            query_name = "SELECT student_name FROM students WHERE student_id = %s LIMIT 1"  
            cursor.execute(query_name, (student_id,))
            
            
            student_name_result = cursor.fetchone()

            if student_name_result:  
                student_name = student_name_result[0]  
                self.student_name_display.config(text=student_name)  
            else:
                self.student_name_display.config(text="Not Found")  

            
            query_courses = "SELECT course FROM students WHERE student_id = %s"  
            cursor.execute(query_courses, (student_id,))
            
            
            course_results = cursor.fetchall()

            
            course_names = []

            if course_results:  
                for row in course_results:
                    course = row[0]  
                    course_names.append(course)  

            
            self.course_name_combobox['values'] = course_names  
            if course_names:  
                self.course_name_combobox.current(0)

            else:
                self.course_name_combobox['values'] = []  

        except mysql.connector.Error as err:
            
            messagebox.showerror("Database Error", f"Error: {err}")
        finally:
            
            cursor.close()
            connection.close()


    def delete_selected_course(self):
        
        selected_index = self.selected_courses_listbox.curselection()

        if selected_index:  
            self.selected_courses_listbox.delete(selected_index)  
        else:
            messagebox.showwarning("Selection Error", "Please select a course to delete.")  


    
    def add_course(self):
        selected_course = self.course_name_combobox.get()  
        if selected_course not in self.selected_courses_listbox.get(0, END):  
            self.selected_courses_listbox.insert(END, selected_course)  

    def back_to_main(self):
        self.root.destroy()  
        new_window = Tk()  
        admit_interface.Admit_Interface(new_window, self.username)



if __name__ == "__main__":
    root = tk.Tk()
    report_generator = Report_Generater(root, "AdminUser")
    root.mainloop()
